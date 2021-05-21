# -*- coding: utf-8 -*-
"""
Created on Fri Nov 20 10:28:43 2020

@author: ovindu wijethunge
"""
from nltk.tokenize import word_tokenize
from func_geting_features import no_sentences, com_length,no_punctuations,check_black_words_list,duplicate_words, find_mail_site_pnumber,calculate_content_comment_similerity
from func_vector_create import vectorize
from func_preprocessing import preprocess,remove_punctuation
from func_find_duplicate_comments import find_duplicate_comments
from func_mail_mob_link_length import link_mob_mail_length_blckword
from func_make_csv import makes_csv, create_file
import numpy as np
from datetime import datetime
import openpyxl
import re
import pandas as pd

# client id = 797659646664-sc1saj14036j4ulm5u49bimpsm1cgitk.apps.googleusercontent.com
# client secret =  NaEFFaU2WXbAnA_jynaH7t26

def mainScript():
    
    wb1 = openpyxl.load_workbook('contents.xlsx')
    wb2 = openpyxl.load_workbook('comments.xlsx') 
    
    df = pd.read_excel('comments.xlsx') 
    li_id = df['comment_id'].tolist()
    
    worksheet1 = wb1.active
    worksheet2 = wb2.active
    number_of_round = (worksheet1.max_row + 1)
    fields_list = []
    for x in range(2,number_of_round):
       # print("round????????????????????????????????????????????????????????????????????????????????????????? ")
        vid = worksheet1.cell(row=2, column=2).value
        main_list = []
        post_date = ""
        for row in worksheet2.iter_rows():
            
            if row[0].value == vid:
                row_list = []
                count = 1
                for cell in row:
                    row_list.append(worksheet2.cell(row=cell.row, column=count).value)
                    post_date = worksheet2.cell(row=cell.row, column=2).value
                    count +=1
                main_list.append(row_list)
                
        
        dataset = []
        comments_date = []
        #classifier = []
        for li in main_list:
            dataset.append(li[3])
            comments_date.append(li[8])
            #classifier.append(li[9])
        content_string = worksheet1.cell(row=2, column=1).value
        dataset.insert(0,content_string)
        N = len(dataset)
        
        def de_emojies(text1):
            regrex_pattern = re.compile(pattern = "["
                u"\U0001F600-\U0001F64F"  # emoticons
                u"\U0001F300-\U0001F5FF"  # symbols & pictographs
                u"\U0001F680-\U0001F6FF"  # transport & map symbols
                u"\U0001F1E0-\U0001F1FF"  # flags (iOS)
                 u"\U00002702-\U000027B0"
                                   u"\U000024C2-\U0001F251"
                                   "]+", flags = re.UNICODE)
            text = regrex_pattern.sub(r' ඉමොජි  ',text1)
          
            return text
        new_dataset = []
        def remove_emojies(dataset):
            d_list = dataset
            for one_list in d_list:
                #sentense = ' '.join(one_list)
                clean_text = de_emojies(one_list)
                new_dataset.append(clean_text)
            
        remove_emojies(dataset)
        
        def calculate_time_gap(val):
           # rep1 = post_date.replace("Z", "+00:00")
            #rep2 = val.replace("Z", "+00:00")
            
            #x = datetime.fromisoformat(rep1).timestamp()
           # y = datetime.fromisoformat(rep2).timestamp()
            #dif = (y-x)/60 # calculate from minitues
            #return dif
            obj1 = datetime.strptime(post_date,'%Y-%m-%dT%H:%M:%SZ')
            obj2 = datetime.strptime(val,'%Y-%m-%dT%H:%M:%SZ')
            time_delta = (obj2-obj1)
            total_seconds = time_delta.total_seconds()
            #minutes = total_seconds/60
            return total_seconds
        
        def generate_ngrams(text):
            
            text = re.sub(r'[#$%&()*+-./:;<=>?\s]', '',text)
            tokens = word_tokenize(text)
           # print(tokens) # ['මචං', 'මැසේජ්', 'වගේ']
            list_comment = []
            for token in tokens:# take a one word at  a time
                n = len(token)
                list_word = []
                for i in range(0,(n-1)):
                    bi = token[i:i+2]
                    list_word.append(bi)
                #print("word bi grams list ",list_word)    
                list_comment.extend(list_word)
        
            #print("comment bi grams ",list_comment)
            return list_comment
        
        # claculate peiods sequence
        def calc(st):
        
            x = ''
            y = ''
            for i in st:
                y = x
                x = i
                if x=='.'and y=='.':
                    return 1
                #print(i)
            return 0    
        
        
        
    
        bi_gram_processed_text = [] #this is a list in list which have bi-gram strings duplicate bi-grams also here
        words_processed_text = [] # this is a list in list which have words tokens strings duplicate words also here
        
        
        for text in new_dataset[:N]:
            text_preprocessed_bi_gram = preprocess(text) # remove stop words and punctuations for make bi grams for find similerity
            bi_grams = generate_ngrams(text_preprocessed_bi_gram)# make a character bigram list for a comment
            bi_gram_processed_text.append(bi_grams)# make a list in list of bigrams for each comments
            
            
            words_processed_text.append(word_tokenize(str(remove_punctuation(text))))
            
        vectors_list = vectorize(bi_gram_processed_text)
        
    #---------------------------------------------------------------------------------------
    
        #center_vector = np.zeros((len(vectors_list[2]))) # initialize for comments center vector
        dynamic_center_vector = []
        for numb in range(1,len(vectors_list)):
            
            center_vector = np.zeros((len(vectors_list[0]))) # initialize for comments center vector
            count = 0
            for i in vectors_list[1:]:
                count +=1
                if count == numb:
                    continue
                else:
                    
                    center_vector = np.add(center_vector,i)
                
            row_center_vector = center_vector/(len(vectors_list)-2)        
            dynamic_center_vector.append(row_center_vector)
        
        
        content_vector = vectors_list[0]
        
        for i in range(1, N):
            cos_simillerity_content_comment = calculate_content_comment_similerity(content_vector,vectors_list[i].tolist())
            cos_simillerity_comment_comment = calculate_content_comment_similerity(dynamic_center_vector[i-1],vectors_list[i].tolist())    
            #word_count =len(words_processed_text[i]) 
            duplicate_word_ratio = duplicate_words(words_processed_text[i])
            no_of_sentences = no_sentences(new_dataset[i])
            #length_of_comment = com_length(new_dataset[i])
            num_of_punctuations = no_punctuations(new_dataset[i])
            is_period_sequence = calc(new_dataset[i])
            #stop_word_ratio = count_stop_word(words_processed_text[i])
            post_coment_gap = calculate_time_gap(comments_date[i-1])
            #is_black_word = check_black_words_list(words_processed_text[i])
            #link_mail_pnumber = find_mail_site_pnumber(new_dataset[i])
            comment_duplication = find_duplicate_comments(main_list , i-1)
            link_mob_mail_word_length = link_mob_mail_length_blckword(words_processed_text[i])
            
            #classifier_val = classifier[i-1]
            
            
            print(' ')
            print(new_dataset[i])
            print("cos_simillerity_content_comment ",cos_simillerity_content_comment)
            print("cos_simillerity_comment_comment  ",cos_simillerity_comment_comment)
            print("word_count  ",link_mob_mail_word_length[1])
            print("duplicate_word_ratio  ",duplicate_word_ratio)
            print("no_of_sentences  ",no_of_sentences)
            print("length_of_comment ",link_mob_mail_word_length[0])
            print("num_of_punctuations ",num_of_punctuations)
            print("is period sequence ",is_period_sequence)
            #print("stop_word_ratio ",stop_word_ratio)
            print("post_coment_gap ",post_coment_gap)
            print("black_word_ratio ",link_mob_mail_word_length[-1])
            print("is_link ",link_mob_mail_word_length[2])
            print("is_youtube_link ",link_mob_mail_word_length[3])
            print("is_number ",link_mob_mail_word_length[4])
            #print("is_mail ",link_mob_mail_word_length[5])
            print("comment_duplication  ",comment_duplication)
    
    
            stringList = makes_csv(li_id[i-1], cos_simillerity_content_comment,cos_simillerity_comment_comment,link_mob_mail_word_length[1],duplicate_word_ratio,no_of_sentences,link_mob_mail_word_length[0],num_of_punctuations,is_period_sequence,post_coment_gap,link_mob_mail_word_length[-1],link_mob_mail_word_length[2],link_mob_mail_word_length[3],link_mob_mail_word_length[4],comment_duplication)
            fields_list.append(stringList)
            
            
        create_file(fields_list)       
           
        
#mainScript()       
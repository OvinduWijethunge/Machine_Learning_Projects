# -*- coding: utf-8 -*-
"""
Created on Sun Jan 31 14:55:31 2021

@author: Ovindu Wijethunge
"""

from __future__ import unicode_literals
import os
import pickle
import google.oauth2.credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
import csv
from nltk.tokenize import word_tokenize
import re
import numpy as np
import xlsxwriter 
from googleapiclient.discovery import build


import youtube_dl
from youtube_dl import YoutubeDL
import speech_recognition as sr 
from pydub import AudioSegment
from pydub.silence import split_on_silence
from pathlib import Path

import openpyxl
import json

from pytube import YouTube


from moviepy.editor import *

import os, shutil
import pandas as pd



CLIENT_SECRETS_FILE = "client_secret.json" # for more information   https://python.gotrained.com/youtube-api-extracting-comments/
SCOPES = ['https://www.googleapis.com/auth/youtube.force-ssl']
API_SERVICE_NAME = 'youtube'
API_VERSION = 'v3'

list_date_title = []
def get_video_date_hate(vid):
    api_key = 'AIzaSyCUUaAbUbpJqD-BE74quAoVFDp-ggd3SLo'
    youtube = build('youtube', 'v3', developerKey=api_key)
    request = youtube.videos().list(

    part = 'snippet',
    id = vid
    )
    response = request.execute()
    video_published_date = "2000-10-25T04:31:22Z"
    video_title = "helloo"
    
    for item in response['items']:
        video_published_date = item['snippet']['publishedAt']
        list_date_title.append(video_published_date)
        video_title = item['snippet']['title']
        list_date_title.append(video_title)
    #return list_date_title

def get_authenticated_service_hate():
    credentials = None
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            credentials = pickle.load(token)
    #  Check if the credentials are invalid or do not exist
    if not credentials or not credentials.valid:
        # Check if the credentials have expired
        if credentials and credentials.expired and credentials.refresh_token:
            credentials.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                CLIENT_SECRETS_FILE, SCOPES)
            credentials = flow.run_console()

        # Save the credentials for the next run
        with open('token.pickle', 'wb') as token:
            pickle.dump(credentials, token)

    return build(API_SERVICE_NAME, API_VERSION, credentials = credentials)




def write_to_excel_hate(d_list):
    
    wb = openpyxl.Workbook()
    dest_filename = 'hamData.xlsx'
    worksheet = wb.active
     
    worksheet['A1']= 'vid' 
    worksheet['B1']= 'vdate'
    worksheet['C1']= 'comment_id'
    worksheet['D1']= 'comment'
    worksheet['E1']= 'channel_id'
    worksheet['F1']= 'video_title'
    worksheet['G1']= 'like_count'
    worksheet['H1']= 'published_date'
    worksheet['I1']= 'updated_date'

    row_count = 2
    for listx in d_list:
        
        list_index = 0
        for j in range(1,len(listx)+1):
            
            worksheet.cell(row=row_count, column=j).value = listx[list_index]    # count , j = row , column
            list_index +=1
        row_count = row_count + 1
    
    wb.save(filename = dest_filename) 
    wb.close()
        
        
def write_to_excel_spam(d_list):
    
    wb = openpyxl.Workbook()
    dest_filename = 'spamData.xlsx'
    worksheet = wb.active
     
    worksheet['A1']= 'vid' 
    worksheet['B1']= 'vdate'
    worksheet['C1']= 'comment_id'
    worksheet['D1']= 'comment'
    worksheet['E1']= 'channel_id'
    worksheet['F1']= 'video_title'
    worksheet['G1']= 'like_count'
    worksheet['H1']= 'published_date'
    worksheet['I1']= 'updated_date'

    row_count = 2
    for listx in d_list:
        
        list_index = 0
        for j in range(1,len(listx)+1):
            
            worksheet.cell(row=row_count, column=j).value = listx[list_index]    # count , j = row , column
            list_index +=1
        row_count = row_count + 1
    
    wb.save(filename = dest_filename) 
    wb.close()        
        
    

def remove_spam_comments(comment_list,spamId_list):
    ham_comments_list = []   
    
    comments = comment_list
    for comment in comments:
        
        if comment[2] in spamId_list:
            
            continue
        
        else:
            ham_comments_list.append(comment)
    return ham_comments_list        

def remove_ham_comments(data_list,spamId_list):
    spam_comments_list = []   
    
    comments = data_list
    for comment in comments:
        
        if comment[2] in spamId_list:
            
            spam_comments_list.append(comment)
        
        else:
            continue
    return spam_comments_list    

 

def get_video_comments(service, **kwargs):
    data_list = []
    data_row = []
    results = service.commentThreads().list(**kwargs).execute()
    print("results ",results)
    while results:
        
        for item in results['items']:
            print(item)
            data_row = []
            print("--------------------------------------------------------------")
            print("item ",item)
            vid = item['snippet']['topLevelComment']['snippet']['videoId']
            data_row.append(vid)
            

            vdate = list_date_title[0]
            data_row.append(vdate)

            
            comment_id = item['snippet']['topLevelComment']['id']
            data_row.append(comment_id)
            
            comment = item['snippet']['topLevelComment']['snippet']['textDisplay']
            
            
            
            text = comment
            
            
            tokens = word_tokenize(str(text))
            length = len(tokens)/2
            pattern = re.compile("[A-Za-z]+")
            count = 0
            for token in tokens: # for a word
                if pattern.fullmatch(token) is not None:
                    count = count + 1
            if count >= length:
                continue
            
            
            
            if length <= 1:
                continue
            #30MRBuXy2RE&list=PLcJsqwQYyc2j4JXT5CKNF_DaR-D0KFdyL
            data_row.append(comment)
            
            channel_id = item['snippet']['topLevelComment']['snippet']['authorChannelId']['value']
            data_row.append(channel_id)
            
            video_title = list_date_title[1]
            data_row.append(video_title)
            
            like_count = item['snippet']['topLevelComment']['snippet']['likeCount']
            data_row.append(like_count)
            
            published_date = item['snippet']['topLevelComment']['snippet']['publishedAt']
            data_row.append(published_date)
            
            updated_date = item['snippet']['topLevelComment']['snippet']['updatedAt']
            data_row.append(updated_date)
            
            data_list.append(data_row)
        # Check if another page exists
        if 'nextPageToken' in results:
            kwargs['pageToken'] = results['nextPageToken']
            results = service.commentThreads().list(**kwargs).execute()
        else:
            break

    return data_list


def get_ham_comments(video_id,spam_list):
    # When running locally, disable OAuthlib's HTTPs verification. When
    # running in production *do not* leave this option enabled.
    os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '1'
    service = get_authenticated_service_hate()
    videoId = video_id
    get_video_date_hate(videoId)
   
    
    data_list = get_video_comments(service, part= 'id,snippet', videoId=videoId, textFormat='plainText')
    ham_data_list = remove_spam_comments(data_list,spam_list)
    spam_data_list = remove_ham_comments(data_list,spam_list)
    #print(data_list)
     #write_to_csv(data_list)
    #print("date is ",date_title)
    write_to_excel_hate(ham_data_list)
    write_to_excel_spam(spam_data_list)
    ham_df =pd.read_excel('hamData.xlsx') 
    ham_df.to_csv('hamDataCsv.csv', header=True, index=False, encoding='utf-16')
    
#video_id = 'BHhl75a5bks'
#id_list  = ['Ugy41He-E7NJsT2dfDl4AaABAg','UgxgZPVRvPhTX_SyDEx4AaABAg','Ugz_wx-oufFe_q9-Wfh4AaABAg']  

#get_ham_comments(video_id,id_list)
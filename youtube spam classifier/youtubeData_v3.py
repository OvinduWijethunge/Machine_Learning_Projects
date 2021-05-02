# -*- coding: utf-8 -*-
"""
Created on Thu Jan 14 21:22:19 2021

@author: Acer
"""
from __future__ import unicode_literals
import os
import pickle
import google.oauth2.credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
import csv
from nltk.tokenize import word_tokenize
import re
import numpy as np
import xlsxwriter 
from googleapiclient.discovery import build


import youtube_dl
from youtube_dl import YoutubeDL
import speech_recognition as sr 
from pydub import AudioSegment
from pydub.silence import split_on_silence
from pathlib import Path

import openpyxl
import json

from pytube import YouTube


from moviepy.editor import *

import os, shutil





CLIENT_SECRETS_FILE = "client_secret.json" # for more information   https://python.gotrained.com/youtube-api-extracting-comments/
SCOPES = ['https://www.googleapis.com/auth/youtube.force-ssl']
API_SERVICE_NAME = 'youtube'
API_VERSION = 'v3'

list_date_title = []
def get_video_date(vid):
    api_key = 'AIzaSyCUUaAbUbpJqD-BE74quAoVFDp-ggd3SLo'
    youtube = build('youtube', 'v3', developerKey=api_key)
    request = youtube.videos().list(

    part = 'snippet',
    id = vid
    )
    response = request.execute()
    video_published_date = "2000-10-25T04:31:22Z"
    video_title = "helloo"
    
    for item in response['items']:
        video_published_date = item['snippet']['publishedAt']
        list_date_title.append(video_published_date)
        video_title = item['snippet']['title']
        list_date_title.append(video_title)
    #return list_date_title

def get_authenticated_service():
    credentials = None
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            credentials = pickle.load(token)
    #  Check if the credentials are invalid or do not exist
    if not credentials or not credentials.valid:
        # Check if the credentials have expired
        if credentials and credentials.expired and credentials.refresh_token:
            credentials.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                CLIENT_SECRETS_FILE, SCOPES)
            credentials = flow.run_console()

        # Save the credentials for the next run
        with open('token.pickle', 'wb') as token:
            pickle.dump(credentials, token)

    return build(API_SERVICE_NAME, API_VERSION, credentials = credentials)



def write_to_excel_content(content_list):
    if os.path.exists('./testcontents.xlsx'):
        wb = openpyxl.load_workbook('testcontents.xlsx')
        worksheet = wb.active
        for listx in content_list:
            
            worksheet.append(listx)
            
        wb.save('testcontents.xlsx')
        wb.close()
    else:
        wb = openpyxl.Workbook()
        dest_filename = 'testcontents.xlsx'
        worksheet = wb.active
         
        worksheet['A1']= 'content' 
        worksheet['B1']= 'vid'
       
        row_count = 2
        for listx in content_list:
            
            list_index = 0
            for j in range(1,len(listx)+1):
                
                worksheet.cell(row=row_count, column=j).value = listx[list_index]    # count , j = row , column
                list_index +=1
            row_count = row_count + 1
        
        wb.save(filename = dest_filename) 
        wb.close()
    
    


def write_to_excel(d_list):
    if os.path.exists('./testcomments.xlsx'):
        wb = openpyxl.load_workbook('testcomments.xlsx')
        worksheet = wb.active
        for listx in d_list:
            
            worksheet.append(listx)
            
        wb.save('testcomments.xlsx')
        wb.close()
    else:
        wb = openpyxl.Workbook()
        dest_filename = 'testcomments.xlsx'
        worksheet = wb.active
         
        worksheet['A1']= 'vid' 
        worksheet['B1']= 'vdate'
        worksheet['C1']= 'comment_id'
        worksheet['D1']= 'comment'
        worksheet['E1']= 'channel_id'
        worksheet['F1']= 'video_title'
        worksheet['G1']= 'like_count'
        worksheet['H1']= 'published_date'
        worksheet['I1']= 'updated_date'
    
        row_count = 2
        for listx in d_list:
            
            list_index = 0
            for j in range(1,len(listx)+1):
                
                worksheet.cell(row=row_count, column=j).value = listx[list_index]    # count , j = row , column
                list_index +=1
            row_count = row_count + 1
        
        wb.save(filename = dest_filename) 
        wb.close()
    

    

def write_to_csv(data_list):
    list_big = []
    list_add = []
    fields = ["vid","comment_id","comment","channel_id","video_title","like_count","published_date","updated_date"]
    # print("coments ",comments)
    
    for i in range(0,len(data_list)):
        list_add = data_list[i]
        stringList1 = [str(x) for x in list_add]
        print(stringList1)
        list_big.append(stringList1)
    print("list big ",list_big)   

    filename = "ManiyaApiDataCsv.csv"
    with open(filename, 'w', newline='', encoding='utf-16') as csvfile:
        csvwriter = csv.writer(csvfile, delimiter='\t')
        csvwriter.writerow(fields)
        csvwriter.writerows(list_big)
        
#      output file  2mq3pCIN7A0  
#       sinahala    4edY1pNRmnI ,kanna = F4OKB86r9BQ  chathura = loO6ws2X50Y 
#       mandha pama RsEoXcrnSxg  , pandhama GNgmUYAsB-M yohani Y60Z56MDIv4
# ratta  RVHZFAFW6rA
# maniya GT9-SQO3aH4


def convert_audio_to_text(filepath, chunksize=30000):
    
    sound = AudioSegment.from_mp3(filepath)

    #1: split file into 60s chunks
    def divide_chunks(sound, chunksize):
        # looping till length l
        for i in range(0, len(sound), chunksize):
            yield sound[i:i + chunksize]
    chunks = list(divide_chunks(sound, chunksize))
    print(f"{len(chunks)} chunks of {chunksize/1000}s each")

    r = sr.Recognizer()
    #2: per chunk, save to wav, then read and run through recognize_google() en-US  si-LK
    string_index = {}
    try:
        
        
        for index,chunk in enumerate(chunks):
            #TODO io.BytesIO()
            chunk.export('E:/Machine_Learning_Projects/youtube spam classifier/alfa.wav', format='wav')
            with sr.AudioFile('E:/Machine_Learning_Projects/youtube spam classifier/alfa.wav') as source:
                audio = r.record(source)
            #s = r.recognize_google(audio, language="en-US") #, key=API_KEY) --- my key results in broken pipe
            response = r.recognize_google(audio, language="si-LK")
            #response = json.dumps(s, ensure_ascii=False).encode('utf8')
            print("val is ",response)
            print(index)
            string_index[index] = response
            #break
        return string_index
    except sr.UnknownValueError:
        print("sr.UnknownValueError ")
    return string_index



def get_video_content(videoId):
    
    url = 'https://www.youtube.com/watch?v='+videoId
    output = "mp3"
    print("Converting...")

    mp4 = YouTube(url).streams.get_highest_resolution().download()
    mp3 = mp4.split(".mp4", 1)[0] + f".{output}"

    video_clip = VideoFileClip(mp4)
    audio_clip = video_clip.audio
    audio_clip.write_audiofile(mp3)

    audio_clip.close()
    video_clip.close()
    print("your mp3 is ",mp3)
    os.rename(mp3,'nova.mp3')
    os.remove(mp4)
    text = convert_audio_to_text('nova.mp3')
    os.remove('alfa.wav')
    
    len_of_text = len(text)
    content_string = ""
    for index in range(0,len_of_text):
        content_string = content_string +" "+ text[index]
    print("your text is  ... ",content_string)
    
    os.remove('nova.mp3')
    return content_string
    #shutil.move(mp3, r"E:\4th year project\scripts\test_scripts")  # Replace this with your own output directory
    # https://www.youtube.com/watch?v=dFEVVgEHtZI
    
    


def de_emojies(text1):
    regrex_pattern = re.compile(pattern = "["
        u"\U0001F600-\U0001F64F"  # emoticons
        u"\U0001F300-\U0001F5FF"  # symbols & pictographs
        u"\U0001F680-\U0001F6FF"  # transport & map symbols
        u"\U0001F1E0-\U0001F1FF"  # flags (iOS)
         u"\U00002702-\U000027B0"
                           u"\U000024C2-\U0001F251"
                           "]+", flags = re.UNICODE)
    text = regrex_pattern.sub(r'',text1)
  
    return text
 

def get_video_comments(service, **kwargs):
    data_list = []
    data_row = []
    results = service.commentThreads().list(**kwargs).execute()
    print("results ",results)
    while results:
        
        for item in results['items']:
            print(item)
            data_row = []
            print("--------------------------------------------------------------")
            print("item ",item)
            vid = item['snippet']['topLevelComment']['snippet']['videoId']
            data_row.append(vid)
            

            vdate = list_date_title[0]
            data_row.append(vdate)

            
            comment_id = item['snippet']['topLevelComment']['id']
            data_row.append(comment_id)
            
            comment = item['snippet']['topLevelComment']['snippet']['textDisplay']
            
            
            symbols = "!\"#$%&()*+-./:;<=>?@[\]^_`{|}~\n"
            text = comment
            for i in range(len(symbols)): # for a comment
                
    
                text = np.char.replace(text, symbols[i], ' ')
                text = np.char.replace(text, "  ", " ")
                text = np.char.replace(text, ',', '')
            
            tokens = word_tokenize(str(text))
            length = len(tokens)/2
            pattern = re.compile("[A-Za-z]+")
            count = 0
            for token in tokens: # for a word
                if pattern.fullmatch(token) is not None:
                    count = count + 1
            if count >= length:
                continue
            
            re_text = de_emojies(comment)
            
            if length <= 1:
                continue
            #30MRBuXy2RE&list=PLcJsqwQYyc2j4JXT5CKNF_DaR-D0KFdyL
            data_row.append(re_text)
            
            channel_id = item['snippet']['topLevelComment']['snippet']['authorChannelId']['value']
            data_row.append(channel_id)
            
            video_title = list_date_title[1]
            data_row.append(video_title)
            
            like_count = item['snippet']['topLevelComment']['snippet']['likeCount']
            data_row.append(like_count)
            
            published_date = item['snippet']['topLevelComment']['snippet']['publishedAt']
            data_row.append(published_date)
            
            updated_date = item['snippet']['topLevelComment']['snippet']['updatedAt']
            data_row.append(updated_date)
            
            data_list.append(data_row)
        # Check if another page exists
        if 'nextPageToken' in results:
            kwargs['pageToken'] = results['nextPageToken']
            results = service.commentThreads().list(**kwargs).execute()
        else:
            break

    return data_list


if __name__ == '__main__':
    # When running locally, disable OAuthlib's HTTPs verification. When
    # running in production *do not* leave this option enabled.
    os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '1'
    service = get_authenticated_service()
    videoId = input('Enter Video id : ') # video id here (the video id of https://www.youtube.com/watch?v=vedLpKXzZqE -> is vedLpKXzZqE)
    date_title = get_video_date(videoId) 
    data_list = get_video_comments(service, part= 'id,snippet', videoId=videoId, textFormat='plainText')
    print(data_list)
    video_content = get_video_content(videoId)  # 0FXKASB1Bd0 _VLjevnS8lw  loO6ws2X50Y
                                                   # #  BW38guk_fQQ  Rjb9sLL0LZI lUukWG4Fqow
    write_to_excel(data_list)
    content_list = [[video_content,videoId]]
    write_to_excel_content(content_list)
  
#print(len(data_list))

#print("my comments ",comments)
# -*- coding: utf-8 -*-
"""
Created on Tue Nov 10 23:16:22 2020

@author: Acer
"""
import os
import openpyxl

def write_to_excel_content(content_list):
    
    if os.path.exists('./alfa.xlsx'):
        wb = openpyxl.load_workbook('alfa.xlsx')
        worksheet = wb.active
        for listx in content_list:
            
            worksheet.append(listx)
            
        wb.save('alfa.xlsx')
        wb.close()
    else:
        wb = openpyxl.Workbook()
        dest_filename = 'alfa.xlsx'
        worksheet = wb.active
         
        worksheet['A1']= 'vid' 
        worksheet['B1']= 'content'
        worksheet['C1']= 'comment_id'
    
        row_count = 2
        for listx in content_list:
            
            list_index = 0
            for j in range(1,len(listx)+1):
                
                worksheet.cell(row=row_count, column=j).value = listx[list_index]    # count , j = row , column
                list_index +=1
            row_count = row_count + 1
        
        wb.save(filename = dest_filename) 
        
content_list = [[34,89,76],[11,22,33],[44,55,66]]
write_to_excel_content(content_list)  
print("ok.....")          
        
# =============================================================================
# =============================================================================
#     wb = Workbook()
#     dest_filename = 'alfa.xlsx'
#     # grab the active worksheet
#     worksheet = wb.active
#   
# 
#     worksheet['A1']= 'vid' 
#     worksheet['B1']= 'content'
#     worksheet['C1']= 'comment_id'
#     
#     row_count = 2
#     for listx in content_list:
#         list_index = 0
#         for j in range(1,len(listx)+1):
#             
#             #sheet.cell(row=2, column=2).value = 2
#             worksheet.cell(row=row_count, column=j).value = listx[list_index]    # count , j = row , column
#             list_index +=1
#         row_count = row_count + 1
#         
#     wb.save(filename = dest_filename) 
# =============================================================================
# =============================================================================
    
